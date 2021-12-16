#-------------------------------------------------------------------------------
# Name:        JiraExportTestCasesParsing
# Purpose:     JiraExportTestCasesParsing
#
# Author:      Raju Surapuraju
#
# Created:     30-11-2021
# Copyright:   (c) Raju Surapuraju 2021
# Licence:     Raju Surapuraju
#-------------------------------------------------------------------------------

import os
import datetime
from configparser import ConfigParser

from collections import OrderedDict
from itertools import islice
from openpyxl import load_workbook
import json

fileDir = os.path.dirname(os.path.realpath('__file__'))
config = ConfigParser()
configFN = "configJiraXLSXParsing.ini"

config.read(configFN)
JiraXLSXFileLocation = config.get('section', 'JiraXLSXFileLocation')
TestCaseFileLocation = config.get('section', 'TestCaseFileLocation')

header = [u'Key', u'Summary', u'Actions', u'Data', u'Expected Results']

dest_filename = TestCaseFileLocation
wb_write = load_workbook(dest_filename)
ws1 = wb_write.active
ws1.title = "Test Casess"
ws1.append(header)


def findAll(string,word):
    all_positions=[]
    next_pos=-1
    while True:
        next_pos=string.find(word,next_pos+1)
        if(next_pos<0):
            break
        all_positions.append(next_pos)
    return all_positions

def countX(lst, x):
    count = 0
    for ele in lst:
        if (ele == x):
            count = count + 1
    return count


def parseTestCases(cell_obj_c1,cell_obj_c2,cell_obj_c3):
    #print(cell_obj_c1.value, end = ": " + "\n")
    #print(cell_obj_c2.value, end = ": " + "\n")
    #print(cell_obj_c3.value, end = ": " + "\n")

    cell_obj_c1 = str(cell_obj_c1.value)
    cell_obj_c2 = str(cell_obj_c2.value)
    cell_obj_c3 = str(cell_obj_c3.value)

    actionsString = 'Actions'
    dataString = 'Data'
    expectedResultsString = 'Expected Results'
    attachmentsString = 'attachments'

    resultActions=findAll(cell_obj_c3,actionsString)
    resultData=findAll(cell_obj_c3,dataString)
    resultExpectedResults=findAll(cell_obj_c3,expectedResultsString)
    attachmentsResults=findAll(cell_obj_c3,attachmentsString)

    actionsCount = len(resultActions)
    dataCount = len(resultData)
    expectedResultsCount = len(resultExpectedResults)
    attachmentsCount = len(attachmentsResults)

    print(resultActions[0])
##        print(resultData[0])
##        print(resultExpectedResults[0])
##        print(attachmentsResults[0])
##        print()

    for loop in range (0, actionsCount):
        resultActionsStart = resultActions[loop]
        resultActionsEnd = resultData[loop]

        resultDataStart = resultData[loop]
        resultDataEnd = resultExpectedResults[loop]

        resultExpectedResultsStart = resultExpectedResults[loop]
        resultExpectedResultsEnd = attachmentsResults[loop]

        actionsStringFinal = cell_obj_c3[resultActionsStart+10:resultActionsEnd-4]
        dataStringFinal = cell_obj_c3[resultDataStart+7:resultDataEnd-3]
        ExpectedResultsStringFinal = cell_obj_c3[resultExpectedResultsStart+19:resultExpectedResultsEnd-4]

        print(actionsStringFinal)
        print(dataStringFinal)
        print(ExpectedResultsStringFinal)
        print()
        print()
        writeTestCases(cell_obj_c1,cell_obj_c2,actionsStringFinal,dataStringFinal,ExpectedResultsStringFinal)

def writeTestCases(cell_obj_c1,cell_obj_c2,actionsStringFinal,dataStringFinal,ExpectedResultsStringFinal):

    #xlsxRow_Write = ["["+"u'" + cell_obj_c1 + "', " + "u'" + cell_obj_c2 + "', " + "u'" + actionsStringFinal + "', " + "u'" + dataStringFinal + "', " + "u'" + ExpectedResultsStringFinal + "'" +"]"]


    xlsxRow_Write = ["'" + cell_obj_c1 +"', "+ "'" + cell_obj_c2 +"', "+ "'" + actionsStringFinal +"', "+ "'" + dataStringFinal + "', " + "'" + ExpectedResultsStringFinal + "'"]


    #for row in new_data:
    ws1.append([cell_obj_c1,cell_obj_c2,actionsStringFinal,dataStringFinal,ExpectedResultsStringFinal])

    wb_write.save(filename = dest_filename)
    wb_write.close()



def readXLSX():

    # Give the location of the file
    InputFile = (JiraXLSXFileLocation)


    wb = load_workbook(InputFile)
    ws = wb["Sheet1"]
    row_count_max = ws.max_row
    #print(row_count_max)

    for i in range(2, row_count_max+1):
        cell_obj_c1 = ws.cell(row = i, column = 1)
        #print(cell_obj_c1.value, end = ": " + "\n")
        cell_obj_c2 = ws.cell(row = i, column = 2)
        #print(cell_obj_c2.value, end = ": " + "\n")
        cell_obj_c3 = ws.cell(row = i, column = 3)
        #print(cell_obj_c3.value, end = ": " + "\n")

        parseTestCases(cell_obj_c1,cell_obj_c2,cell_obj_c3)


def main():
    readXLSX()


if __name__ == '__main__':
    main()

