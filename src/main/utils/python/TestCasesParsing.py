#-------------------------------------------------------------------------------
# Name:        JiraExportTestCasesParsing
# Purpose:
#
# Author:      Raju Surapuraju
#
# Created:     30-11-2021
# Copyright:   (c) Raju Surapuraju 2021
# Licence:     Raju Surapuraju
#-------------------------------------------------------------------------------

import os
from configparser import ConfigParser
import datetime
import openpyxl

fileDir = os.path.dirname(os.path.realpath('__file__'))
config = ConfigParser()
configFN = "configJiraXLSXParsing.ini"

config.read(configFN)
JiraXLSXFileLocation = config.get('section', 'JiraXLSXFileLocation')
TestCaseFileLocation = config.get('section', 'TestCaseFileLocation')

#def writeTestCases():



def parseXLSX():

    # Give the location of the file
    path = (JiraXLSXFileLocation)

    # To open Workbook

    wbObj = openpyxl.load_workbook(path)
    sheetObj = wbObj.active
    #sheetObj=sheetObj.get_sheet_by_name('Sheet1')
    maxRow = sheetObj.max_row
    print(maxRow)
    #maxRow = maxRow + 1

    for currRow in range(maxRow):
        key = sheetObj.cell(row = currRow, column = 1)

        #summary = (sheetObj.cell_value(currRow, 1))
        #testSteps = (sheetObj.cell_value(currRow, 2))


        print(key)
        #print(summary)
        #print(testSteps)




def main():
    parseXLSX()

if __name__ == '__main__':
    main()