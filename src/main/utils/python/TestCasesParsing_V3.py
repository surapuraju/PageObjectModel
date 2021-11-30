#-------------------------------------------------------------------------------
# Name:        JiraExportTestCasesParsing
# Purpose:     JiraExportTestCasesParsing
#
# Author:      Raju Surapuraju
#
# Created:     30-11-2021
# Copyright:   (c) Raju Surapuraju 2021
# Licence:     Raju Surapuraju
#-------------------------------------------------------------------------------

import os
import datetime
from configparser import ConfigParser

from collections import OrderedDict
from itertools import islice
from openpyxl import load_workbook
import json

fileDir = os.path.dirname(os.path.realpath('__file__'))
config = ConfigParser()
configFN = "configJiraXLSXParsing.ini"

config.read(configFN)
JiraXLSXFileLocation = config.get('section', 'JiraXLSXFileLocation')
TestCaseFileLocation = config.get('section', 'TestCaseFileLocation')


def findAll(string,word):
    all_positions=[]
    next_pos=-1
    while True:
        next_pos=string.find(word,next_pos+1)
        if(next_pos<0):
            break
        all_positions.append(next_pos)
    return all_positions

def countX(lst, x):
    count = 0
    for ele in lst:
        if (ele == x):
            count = count + 1
    return count


def writeTestCases(cell_obj_c1,cell_obj_c2,cell_obj_c3):
        print(cell_obj_c1.value, end = ": " + "\n")
        print(cell_obj_c2.value, end = ": " + "\n")
        print(cell_obj_c3.value, end = ": " + "\n")

        #[{"id":239891,"index":1,"fields":{"Actions":"Sign on to APC from KPIM", "Data":"","Expected Results":"Home (Welcone) page is displayed"},"attachments":[]}

        #{"Actions":"Sign on to APC from KPIM", "Data":"","Expected Results":"Home (Welcone) page is displayed"},"attachments":[]}
        #{ "name":"John", "age":30, "city":"New York"}

        cell_obj_c3 = str(cell_obj_c3.value)

        actionsString = 'Actions'

        result1=findAll(cell_obj_c3,actionsString)

        #result=findAll('[{"id":239891,"index":1,"fields":{"Actions":"Sign on to APC from KPIM", "Data":"","Expected Results":"Home (Welcone) page is displayed"},"attachments":[]},{"id":239892,"index":2,"fields":{"Actions":"Sign on to APC from KPIM", "Data":"","Expected Results":"Home (Welcone) page is displayed"},"attachments":[]},{"id":239893,"index":3,"fields":{"Actions":"Sign on to APC from KPIM", "Data":"","Expected Results":"Home (Welcone) page is displayed"},"attachments":[]}]',actionsString)

        #print(result)
        print(len(result1))

        #print('{} has occurred {} times'.format(actionsString, countX(result1, actionsString)))

        #cell_obj_c3 = "test test test test"



        #print (cell_obj_c3.find_all(actionsString))

        #x = str(cell_obj_c3).split(actionsString)

        #print(x)

        #countActions = str(cell_obj_c3).count(actionsString)

        #print(countActions)

##        cell_obj_c3 = '{"Actions":"Sign on to APC from KPIM", "Data":"","Expected Results":"Home (Welcone) page is displayed"}'
##
##        cell_obj_c3 = json.loads(cell_obj_c3)
##        print(cell_obj_c3["Actions"])
##        print(cell_obj_c3["Data"])
##        print(cell_obj_c3["Expected Results"])


        #Get JSON structure length by index
        #Extract the following from JSON structure
        #Action
        #Data
        #Expexted Results
        #Attachments


def parseXLSX():

    # Give the location of the file
    InputFile = (JiraXLSXFileLocation)


    wb = load_workbook(InputFile)
    ws = wb["Sheet1"]
    row_count_max = ws.max_row
    #print(row_count_max)

    for i in range(2, row_count_max):
        cell_obj_c1 = ws.cell(row = i, column = 1)
        #print(cell_obj_c1.value, end = ": " + "\n")
        cell_obj_c2 = ws.cell(row = i, column = 2)
        #print(cell_obj_c2.value, end = ": " + "\n")
        cell_obj_c3 = ws.cell(row = i, column = 3)
        #print(cell_obj_c3.value, end = ": " + "\n")

        writeTestCases(cell_obj_c1,cell_obj_c2,cell_obj_c3)


def main():
    parseXLSX()

if __name__ == '__main__':
    main()

