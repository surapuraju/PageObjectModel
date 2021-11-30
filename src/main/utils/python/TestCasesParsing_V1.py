#-------------------------------------------------------------------------------
# Name:        JiraExportTestCasesParsing
# Purpose:     JiraExportTestCasesParsing
#
# Author:      Raju Surapuraju
#
# Created:     30-11-2021
# Copyright:   (c) Raju Surapuraju 2021
# Licence:     Raju Surapuraju
#-------------------------------------------------------------------------------

import os
import datetime
from configparser import ConfigParser

from collections import OrderedDict
from itertools import islice
from openpyxl import load_workbook
import json

fileDir = os.path.dirname(os.path.realpath('__file__'))
config = ConfigParser()
configFN = "configJiraXLSXParsing.ini"

config.read(configFN)
JiraXLSXFileLocation = config.get('section', 'JiraXLSXFileLocation')
TestCaseFileLocation = config.get('section', 'TestCaseFileLocation')


def serialize_sets(obj):
    if isinstance(obj, set):
        return list(obj)

    return obj


def writeTestCases(cell_obj_c1,cell_obj_c2,cell_obj_c3):
        #print(cell_obj_c1.value, end = ": " + "\n")
        #print(cell_obj_c2.value, end = ": " + "\n")
        #print(cell_obj_c3.value, end = ": " + "\n")

        #[{"id":239891,"index":1,"fields":{"Actions":"Sign on to APC from KPIM", "Data":"","Expected Results":"Home (Welcone) page is displayed"},"attachments":[]}

        #data = []

        cell_obj_c3 = cell_obj_c3.value

        cell_obj_c3 = json.dumps(cell_obj_c3)


        cell_obj_c3 = json.dumps(cell_obj_c3, default=serialize_sets)
        #print(json_str)


        cell_obj_c3 = json.loads(cell_obj_c3)

        for doc in response['result']:
            print(doc['_id'], doc['total'])

        for data in cell_obj_c3['index']:
            print(data['fields'], data['Actions'])

        #print(cell_obj_c3["id"]["Actions"])

        #testcase_list = []

        #for row in islice(cell_obj_c3.value, 1,):
        #    step = OrderedDict()
        #    step['id'] = row[0]
            #step['actions'] = row[1]
            #step['fields'] = row[1]
        #    testcase_list.append(step)

        #j = json.dumps(testcase_list)

        #print(j)

        #Get JSON structure length by index
        #Extract the following from JSON structure
        #Action
        #Data
        #Expexted Results
        #Attachments


def parseXLSX():

    # Give the location of the file
    InputFile = (JiraXLSXFileLocation)


    wb = load_workbook(InputFile)
    ws = wb["Sheet1"]
    row_count_max = ws.max_row
    #print(row_count_max)

    for i in range(2, row_count_max):
        cell_obj_c1 = ws.cell(row = i, column = 1)
        #print(cell_obj_c1.value, end = ": " + "\n")
        cell_obj_c2 = ws.cell(row = i, column = 2)
        #print(cell_obj_c2.value, end = ": " + "\n")
        cell_obj_c3 = ws.cell(row = i, column = 3)
        #print(cell_obj_c3.value, end = ": " + "\n")

        writeTestCases(cell_obj_c1,cell_obj_c2,cell_obj_c3)


def main():
    parseXLSX()

if __name__ == '__main__':
    main()

